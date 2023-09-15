      
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import schedule
import time


filename = "price.xlsx"

if os.path.exists(filename):
    wb = load_workbook(filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws['C1'] = "筛选:二室,有电梯,普通住宅,300~400万"
    ws['A2'] = "时间"
    ws['B2'] = "北京全部"
    ws['C2'] = "北京"
    ws['D2'] = "海淀"
    ws['E2'] = "朝阳"

urls = {
    "bejing":"https://bj.lianjia.com/ershoufang/ie2sf1l2p4/",
    "bejingall":"https://bj.lianjia.com/ershoufang/",
    "haidian": "https://bj.lianjia.com/ershoufang/haidian/ie2sf1l2p4/",
    "chaoyang": "https://bj.lianjia.com/ershoufang/chaoyang/ie2sf1l2p4/"
}

def getByUrl(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    h2_tag = soup.find("h2", class_="total fl")
    span_tag = h2_tag.find("span").get_text()
    return span_tag

def job():
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    last_row = ws.max_row
    last_date_str = ws.cell(row=last_row, column=1).value
    try:
        last_date = datetime.strptime(last_date_str, "%Y-%m-%d %H:%M:%S").date()
        if last_date == datetime.now().date():
            ws.cell(row=last_row, column=1, value=current_date)
            ws.cell(row=last_row, column=2, value=getByUrl(urls["bejingall"]))
            ws.cell(row=last_row, column=3, value=getByUrl(urls["bejing"]))
            ws.cell(row=last_row, column=4, value=getByUrl(urls["haidian"]))
            ws.cell(row=last_row, column=5, value=getByUrl(urls["chaoyang"]))
        else:
            ws.append([current_date, getByUrl(urls["bejingall"]), getByUrl(urls["bejing"]), getByUrl(urls["haidian"]), getByUrl(urls["chaoyang"])])
    except ValueError:
        ws.append([current_date, getByUrl(urls["bejingall"]), getByUrl(urls["bejing"]), getByUrl(urls["haidian"]), getByUrl(urls["chaoyang"])])
    wb.save(filename)
job()
schedule.every().day.at("18:00").do(job)
while True:
    schedule.run_pending()
    time.sleep(1)

    